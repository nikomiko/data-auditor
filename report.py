"""
report.py — Export CSV/HTML et historisation des audits.
"""
import csv
import io
import json
import os
from datetime import datetime
from jinja2 import Template

REPORTS_DIR = os.path.join(os.path.dirname(__file__), "reports")


def save_history(results: list, summary: dict, config: dict) -> str:
    os.makedirs(REPORTS_DIR, exist_ok=True)
    ts        = datetime.now().strftime("%Y%m%d_%H%M%S")
    name      = config.get("meta", {}).get("name", "audit").replace(" ", "_")
    filename  = f"{ts}_{name}.json"
    path      = os.path.join(REPORTS_DIR, filename)
    payload   = {
        "meta":    {"timestamp": datetime.now().isoformat(), "audit_name": name, "config": config},
        "summary": summary,
        "results": results,
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2, default=str)
    return filename


def list_history() -> list:
    os.makedirs(REPORTS_DIR, exist_ok=True)
    files = sorted([f for f in os.listdir(REPORTS_DIR) if f.endswith(".json")], reverse=True)
    out = []
    for fname in files:
        try:
            with open(os.path.join(REPORTS_DIR, fname), encoding="utf-8") as f:
                d = json.load(f)
            out.append({
                "filename":   fname,
                "timestamp":  d.get("meta", {}).get("timestamp", ""),
                "audit_name": d.get("meta", {}).get("audit_name", ""),
                "summary":    d.get("summary", {}),
            })
        except Exception:
            continue
    return out


def load_history(filename: str) -> dict:
    path = os.path.join(REPORTS_DIR, filename)
    if not os.path.exists(path):
        raise FileNotFoundError(f"Rapport introuvable : {filename}")
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def to_csv(results: list) -> str:
    headers = ["join_key", "type_ecart", "rule_name", "champ",
               "valeur_reference", "valeur_cible", "detail"]
    out = io.StringIO()
    w   = csv.DictWriter(out, fieldnames=headers, extrasaction="ignore")
    w.writeheader()
    w.writerows(results)
    return out.getvalue()


def to_xlsx(results: list, summary: dict, config: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment

    name = config.get("meta", {}).get("name", "Audit")
    wb   = Workbook()

    # Feuille Résumé
    ws_sum = wb.active
    ws_sum.title = "Résumé"
    for row in [
        ("Audit",              name),
        ("Généré le",          datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
        ("Référence (lignes)", summary.get("total_reference", 0)),
        ("Cible (lignes)",     summary.get("total_cible", 0)),
        ("Orphelins A",        summary.get("orphelins_a", 0)),
        ("Orphelins B",        summary.get("orphelins_b", 0)),
        ("Divergents",         summary.get("divergents", 0)),
        ("OK",                 summary.get("ok", 0)),
    ]:
        ws_sum.append(row)
    ws_sum.column_dimensions["A"].width = 22
    ws_sum.column_dimensions["B"].width = 30

    # Feuille Résultats
    ws_res = wb.create_sheet("Résultats")
    ws_res.append(["Clé", "Type", "Règle", "Champ", "Valeur réf.", "Valeur cible", "Détail"])
    hdr_fill = PatternFill("solid", fgColor="2D3748")
    hdr_font = Font(color="FFFFFF", bold=True, size=9)
    for cell in ws_res[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="left")

    FILLS = {
        "ORPHELIN_A": PatternFill("solid", fgColor="FFF5F5"),
        "ORPHELIN_B": PatternFill("solid", fgColor="FFFAF0"),
        "KO":         PatternFill("solid", fgColor="FFFFF0"),
        "OK":         PatternFill("solid", fgColor="F0FFF4"),
    }
    for r in results:
        ws_res.append([r.get(k, "") for k in
            ["join_key", "type_ecart", "rule_name", "champ",
             "valeur_reference", "valeur_cible", "detail"]])
        fill = FILLS.get(r.get("type_ecart", ""))
        if fill:
            for cell in ws_res[ws_res.max_row]:
                cell.fill = fill

    for col, w in zip("ABCDEFG", [28, 12, 20, 20, 18, 18, 40]):
        ws_res.column_dimensions[col].width = w

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def to_html(results: list, summary: dict, config: dict) -> str:
    name = config.get("meta", {}).get("name", "Audit")
    now  = datetime.now().strftime("%d/%m/%Y à %H:%M:%S")
    tpl  = Template("""<!DOCTYPE html>
<html lang="fr"><head><meta charset="UTF-8">
<title>Rapport — {{ name }}</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Segoe UI',sans-serif;background:#f4f6f9;color:#2d3748;padding:2rem}
h1{font-size:1.5rem;margin-bottom:.25rem}
.meta{font-size:.8rem;color:#718096;margin-bottom:1.5rem}
.cards{display:flex;gap:1rem;flex-wrap:wrap;margin-bottom:1.5rem}
.card{background:#fff;border-radius:8px;padding:.75rem 1.25rem;box-shadow:0 1px 3px rgba(0,0,0,.1);text-align:center;min-width:110px}
.card .v{font-size:1.75rem;font-weight:700}.card .l{font-size:.7rem;color:#718096;margin-top:.2rem}
.ca .v{color:#e53e3e}.cb .v{color:#dd6b20}.cd .v{color:#d69e2e}
table{width:100%;border-collapse:collapse;background:#fff;border-radius:8px;overflow:hidden;box-shadow:0 1px 3px rgba(0,0,0,.1)}
th{background:#2d3748;color:#fff;padding:.6rem .9rem;text-align:left;font-size:.72rem;text-transform:uppercase;letter-spacing:.05em}
td{padding:.55rem .9rem;font-size:.82rem;border-bottom:1px solid #e2e8f0}
tr:last-child td{border:none}
.badge{display:inline-block;padding:.15rem .5rem;border-radius:4px;font-size:.68rem;font-weight:700}
.badge-ORPHELIN_A{background:#fff5f5;color:#c53030;border:1px solid #fed7d7}
.badge-ORPHELIN_B{background:#fffaf0;color:#c05621;border:1px solid #feebc8}
.badge-DIVERGENT{background:#fffff0;color:#b7791f;border:1px solid #fefcbf}
.badge-OK{background:#f0fff4;color:#276749;border:1px solid #c6f6d5}
.ref{color:#2b6cb0}.tgt{color:#c05621}
code{font-family:monospace;font-size:.8rem;color:#4a5568}
</style></head><body>
<h1>📊 {{ name }}</h1>
<div class="meta">Généré le {{ now }}</div>
<div class="cards">
  <div class="card"><div class="v">{{ s.total_reference }}</div><div class="l">Réf.</div></div>
  <div class="card"><div class="v">{{ s.total_cible }}</div><div class="l">Cible</div></div>
  <div class="card ca"><div class="v">{{ s.orphelins_a }}</div><div class="l">Orphelins A</div></div>
  <div class="card cb"><div class="v">{{ s.orphelins_b }}</div><div class="l">Orphelins B</div></div>
  <div class="card cd"><div class="v">{{ s.divergents }}</div><div class="l">Divergents</div></div>
</div>
{% if results %}
<table><thead><tr><th>Clé</th><th>Type</th><th>Rule</th><th>Champ</th><th>Réf.</th><th>Cible</th><th>Détail</th></tr></thead>
<tbody>{% for r in results %}
<tr>
  <td><code>{{ r.join_key }}</code></td>
  <td><span class="badge badge-{{ r.type_ecart }}">{{ r.type_ecart }}</span></td>
  <td style="color:#7c3aed;font-size:.75rem">{{ r.rule_name }}</td>
  <td>{{ r.champ }}</td>
  <td class="ref">{{ r.valeur_reference }}</td>
  <td class="tgt">{{ r.valeur_cible }}</td>
  <td>{{ r.detail }}</td>
</tr>{% endfor %}
</tbody></table>
{% else %}
<p style="text-align:center;padding:2rem;color:#718096;font-style:italic">✅ Aucun écart détecté.</p>
{% endif %}
</body></html>""")
    return tpl.render(name=name, now=now, s=summary, results=results)
