"""
report.py — Export CSV/HTML/XLSX et historisation des audits.
"""
import csv
import io
import json
import os
from collections import defaultdict
from datetime import datetime

def _get_reports_dir() -> str:
    """Retourne le répertoire des rapports.

    En mode PyInstaller frozen, les rapports sont stockés à côté du .exe
    (persistant entre les mises à jour), et non dans le répertoire temporaire
    d'extraction (_MEIPASS).
    """
    import sys as _sys
    if getattr(_sys, "frozen", False):
        return os.path.join(os.path.dirname(_sys.executable), "reports")
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), "reports")

REPORTS_DIR = _get_reports_dir()

_BASE_FIELDS = ["join_key", "rule_name", "rule_type",
                "source_field", "target_field", "source_value", "target_value", "detail"]


# ─────────────────────────────────────────────────────────────
#  Historisation
# ─────────────────────────────────────────────────────────────

def save_history(results: list, summary: dict, config: dict,
                 run_label: str = "", started_at: str = "", finished_at: str = "") -> str:
    os.makedirs(REPORTS_DIR, exist_ok=True)
    ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
    name     = config.get("meta", {}).get("name", "audit").replace(" ", "_")
    filename = f"{ts}_{name}.json"
    path     = os.path.join(REPORTS_DIR, filename)
    meta: dict = {
        "timestamp":    datetime.now().isoformat(),
        "audit_name":   name,
        "config":       config,
        "total_results": len(results),
    }
    if run_label:    meta["run_label"]    = run_label
    if started_at:   meta["started_at"]  = started_at
    if finished_at:  meta["finished_at"] = finished_at
    if started_at and finished_at:
        try:
            from datetime import datetime as _dt
            d = (_dt.fromisoformat(finished_at) - _dt.fromisoformat(started_at)).total_seconds()
            meta["duration_s"] = round(d, 1)
        except Exception:
            pass
    payload  = {"meta": meta, "summary": summary, "results": results}
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2, default=str)
    return filename


def list_history() -> list:
    os.makedirs(REPORTS_DIR, exist_ok=True)
    files = sorted([f for f in os.listdir(REPORTS_DIR) if f.endswith(".json")], reverse=True)
    out = []
    for fname in files:
        try:
            with open(os.path.join(REPORTS_DIR, fname), encoding="utf-8") as f:
                d = json.load(f)
            m = d.get("meta", {})
            out.append({
                "filename":      fname,
                "timestamp":     m.get("timestamp", ""),
                "audit_name":    m.get("audit_name", ""),
                "run_label":     m.get("run_label", ""),
                "duration_s":    m.get("duration_s"),
                "total_results": m.get("total_results", len(d.get("results", []))),
                "summary":       d.get("summary", {}),
            })
        except Exception:
            continue
    return out


def load_history(filename: str) -> dict:
    path = os.path.join(REPORTS_DIR, filename)
    if not os.path.exists(path):
        raise FileNotFoundError(f"Rapport introuvable : {filename}")
    with open(path, encoding="utf-8") as f:
        return json.load(f)


# ─────────────────────────────────────────────────────────────
#  CSV — pivot-friendly (une ligne par écart)
# ─────────────────────────────────────────────────────────────

def to_csv(results: list,
           config: dict = None,
           extra_ref: list = None, extra_tgt: list = None,
           ref_rows_map: dict = None, tgt_rows_map: dict = None,
           ref_label: str = "Référence", tgt_label: str = "Cible") -> str:
    """Export CSV plat — une ligne par écart avec colonnes supplémentaires.

    Format adapté aux tableaux croisés dynamiques Excel.
    """
    config       = config or {}
    extra_ref    = extra_ref or []
    extra_tgt    = extra_tgt or []
    ref_rows_map = ref_rows_map or {}
    tgt_rows_map = tgt_rows_map or {}

    cfg_keys   = config.get("join", {}).get("keys", [])
    key_fields = [k.get("source_field", "Clé") for k in cfg_keys] if cfg_keys else ["Clé"]
    other_fields = ["rule_name", "rule_type", "source_field", "target_field", "source_value", "target_value", "detail"]

    headers = key_fields + other_fields + [
        f"{ref_label} \u00b7 {c}" for c in extra_ref
    ] + [
        f"{tgt_label} \u00b7 {c}" for c in extra_tgt
    ]

    out = io.StringIO()
    w   = csv.writer(out)
    w.writerow(headers)
    for r in results:
        key   = r.get("join_key", "")
        kp    = key.split("\u00a7")  # § separator
        row   = [kp[i] if i < len(kp) else "" for i in range(len(key_fields))]
        row  += [r.get(k, "") for k in other_fields]
        row  += [ref_rows_map.get(key, {}).get(c, "") for c in extra_ref]
        row  += [tgt_rows_map.get(key, {}).get(c, "") for c in extra_tgt]
        w.writerow(row)
    return out.getvalue()


# ─────────────────────────────────────────────────────────────
#  XLSX — onglet DATA + onglet PIVOT
# ─────────────────────────────────────────────────────────────

def to_xlsx(results: list, summary: dict, config: dict,
            extra_ref: list = None, extra_tgt: list = None,
            ref_rows_map: dict = None, tgt_rows_map: dict = None,
            ref_label: str = "Référence", tgt_label: str = "Cible",
            ref_fmt: str = "", tgt_fmt: str = "") -> bytes:

    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    extra_ref    = extra_ref or []
    extra_tgt    = extra_tgt or []
    ref_rows_map = ref_rows_map or {}
    tgt_rows_map = tgt_rows_map or {}
    name         = config.get("meta", {}).get("name", "Audit")
    cfg_keys     = config.get("join", {}).get("keys", [])
    key_fields   = [k.get("source_field", "Clé") for k in cfg_keys] if cfg_keys else ["Clé"]
    other_hdrs   = ["Règle", "Type", "Champ src.", "Champ cible", "Valeur src.", "Valeur cible", "Détail"]
    other_fields = ["rule_name", "rule_type", "source_field", "target_field", "source_value", "target_value", "detail"]
    n_key        = len(key_fields)
    wb           = Workbook()

    # ── Onglet DATA ───────────────────────────────────────────
    ws = wb.active
    ws.title = "DATA"

    hdr_labels = (
        key_fields + other_hdrs
        + [f"{ref_label} \u00b7 {c}" for c in extra_ref]
        + [f"{tgt_label} \u00b7 {c}" for c in extra_tgt]
    )

    HDR_FILL  = PatternFill("solid", fgColor="2D3748")
    HDR_FONT  = Font(color="FFFFFF", bold=True, size=9)
    REF_HDR_F = PatternFill("solid", fgColor="1E40AF")
    TGT_HDR_F = PatternFill("solid", fgColor="5B21B6")
    REF_ROW_F = PatternFill("solid", fgColor="EFF6FF")
    TGT_ROW_F = PatternFill("solid", fgColor="F5F3FF")

    ws.append(hdr_labels)
    n_base = n_key + len(other_fields)
    n_ref  = len(extra_ref)
    n_tgt  = len(extra_tgt)
    for i, cell in enumerate(ws[1]):
        if i < n_base:
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
        elif i < n_base + n_ref:
            cell.fill = REF_HDR_F
            cell.font = HDR_FONT
        else:
            cell.fill = TGT_HDR_F
            cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="left")

    ROW_FILLS = {
        "_ko": PatternFill("solid", fgColor="FEE5E5"),  # Orphelins
        "ko":  PatternFill("solid", fgColor="FFFBEA"),  # Règle incoherence
        "ok":  PatternFill("solid", fgColor="ECFDF5"),  # Règle coherence
        "_ok": PatternFill("solid", fgColor="EFF6FF"),  # Clé OK
    }

    for r in results:
        key  = r.get("join_key", "")
        rt   = r.get("rule_type", "")
        kp   = key.split("\u00a7")  # § separator
        row  = [kp[i] if i < len(kp) else "" for i in range(n_key)]
        row += [r.get(k, "") for k in other_fields]
        row += [ref_rows_map.get(key, {}).get(c, "") for c in extra_ref]
        row += [tgt_rows_map.get(key, {}).get(c, "") for c in extra_tgt]
        ws.append(row)
        ri = ws.max_row
        base_fill = ROW_FILLS.get(rt)
        for i, cell in enumerate(ws[ri]):
            if n_base <= i < n_base + n_ref:
                cell.fill = REF_ROW_F
            elif n_base + n_ref <= i:
                cell.fill = TGT_ROW_F
            elif base_fill:
                cell.fill = base_fill

    # Table avec autofiltre
    if ws.max_row > 1:
        last_col = get_column_letter(len(hdr_labels))
        tbl = Table(displayName="DATA", ref=f"A1:{last_col}{ws.max_row}")
        tbl.tableStyleInfo = TableStyleInfo(
            name="TableStyleLight9", showRowStripes=True)
        ws.add_table(tbl)

    ws.freeze_panes = "A2"
    col_widths = [20] * n_key + [12, 20, 20, 18, 18, 35] + [20] * (n_ref + n_tgt)
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()



def _build_pivot_sheet(ws, results, summary, config, audit_name):
    """Feuille PIVOT : résumé par règle et par type d'écart."""
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    def cell(row, col, value="", bold=False, size=9, fg="000000", bg=None,
             align="left", wrap=False):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(bold=bold, size=size, color=fg)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        return c

    TITLE_BG = "1E3A5F"
    H2_PRES  = "3B5998"
    H2_RULE  = "2B6CB0"
    H2_SUM   = "374151"
    WHITE    = "FFFFFF"

    def title_row(row, text, ncols=5):
        for c in range(1, ncols + 1):
            ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor=TITLE_BG)
        cell(row, 1, text, bold=True, size=10, fg=WHITE, bg=TITLE_BG)
        ws.row_dimensions[row].height = 20

    def hdr_row(row, labels, bg):
        for i, lbl in enumerate(labels, 1):
            cell(row, i, lbl, bold=True, size=9, fg=WHITE, bg=bg)
        ws.row_dimensions[row].height = 18

    # Titre
    cell(1, 1, f"Rapport d'audit \u2014 {audit_name}", bold=True, size=14)
    cell(2, 1, f"G\u00e9n\u00e9r\u00e9 le {datetime.now().strftime('%d/%m/%Y \u00e0 %H:%M:%S')}",
         size=9, fg="718096")
    ws.row_dimensions[1].height = 24

    row = 4

    # Résumé global
    title_row(row, "R\u00e9capitulatif", 2); row += 1
    hdr_row(row, ["Indicateur", "Valeur"], H2_SUM); row += 1
    for lbl, key in [
        ("R\u00e9f\u00e9rence (lignes)",  "total_reference"),
        ("Cible (lignes)",                "total_cible"),
        ("Orphelins source (A)",          "orphelins_a"),
        ("Orphelins cible  (B)",          "orphelins_b"),
        ("Contr\u00f4les KO",             "divergents"),
        ("Contr\u00f4les OK",             "ok"),
    ]:
        cell(row, 1, lbl, bold=True)
        cell(row, 2, summary.get(key, 0), align="right")
        row += 1
    row += 1

    # Présence
    oa_keys = {r["join_key"] for r in results if r.get("rule_name") == "Source uniq."}
    ob_keys = {r["join_key"] for r in results if r.get("rule_name") == "Cible uniq."}

    title_row(row, "Contr\u00f4les de pr\u00e9sence", 3); row += 1
    hdr_row(row, ["Type", "Description", "Nb cl\u00e9s"], H2_PRES); row += 1
    for rn, desc, keys in [
        ("Source uniq.", "Pr\u00e9sent en r\u00e9f\u00e9rence, absent de la cible", oa_keys),
        ("Cible uniq.", "Absent de la r\u00e9f\u00e9rence, pr\u00e9sent en cible",  ob_keys),
    ]:
        cell(row, 1, rn)
        cell(row, 2, desc, wrap=True)
        cell(row, 3, len(keys), align="right")
        row += 1
    row += 1

    # Par règle
    rule_stats = defaultdict(lambda: defaultdict(
        lambda: {"keys": set(), "ecarts": 0, "champs": set()}
    ))
    for r in results:
        rt = r.get("rule_type"); rn = r.get("rule_name")
        if rt in ("ko", "ok") and rn:
            s = rule_stats[rn][rt]
            s["keys"].add(r.get("join_key", ""))
            s["ecarts"] += 1
            ch = r.get("source_field", "")
            if ch:
                s["champs"].add(ch)

    rule_order = [r["name"] for r in config.get("rules", [])] or sorted(rule_stats)

    title_row(row, "Contr\u00f4les par r\u00e8gle", 5); row += 1
    hdr_row(row, ["R\u00e8gle", "Type",
                  "Nb cl\u00e9s touch\u00e9es", "Nb \u00e9carts",
                  "Champs concern\u00e9s"], H2_RULE)
    row += 1

    for rname in rule_order:
        if rname not in rule_stats:
            continue
        for rt in ("ko", "ok"):
            s = rule_stats[rname].get(rt)
            if not s or not s["ecarts"]:
                continue
            cell(row, 1, rname)
            cell(row, 2, rt.upper())
            cell(row, 3, len(s["keys"]),  align="right")
            cell(row, 4, s["ecarts"],      align="right")
            cell(row, 5, "; ".join(sorted(s["champs"])), wrap=True)
            row += 1

    for col, w in [(1, 30), (2, 14), (3, 20), (4, 12), (5, 50)]:
        ws.column_dimensions[get_column_letter(col)].width = w


# ─────────────────────────────────────────────────────────────
#  HTML — vue courante avec filtres dynamiques JS
# ─────────────────────────────────────────────────────────────

_HTML_JS = r"""
const ALL=__ALL__;
const COL_ORDER=__COL_ORDER__;
const KEY_FIELDS=__KEY_FIELDS__;
let aR=new Set(ALL.flatMap(r=>r.ecarts.map(e=>e.rule_name)));
let sC=null,sD=1;
const BADGE_STYLE={
  '_ko':'background:#fee2e2;color:#b91c1c;border:1px solid #fecaca',
  'ko' :'background:#fef9c3;color:#a16207;border:1px solid #fef08a',
  'ok' :'background:#dcfce7;color:#15803d;border:1px solid #bbf7d0',
  '_ok':'background:#eff6ff;color:#1d4ed8;border:1px solid #bfdbfe',
};
function initCounts(){
  const cr={};
  for(const r of ALL)for(const e of r.ecarts)cr[e.rule_name]=(cr[e.rule_name]||0)+1;
  document.querySelectorAll('[data-k="rule"]').forEach(b=>{
    const sp=b.querySelector('.chip-c');
    if(sp)sp.textContent=(cr[b.dataset.v]||0).toLocaleString('fr-FR');
  });
}
function toggleChip(b){
  b.classList.toggle('on');
  if(b.classList.contains('on'))aR.add(b.dataset.v);else aR.delete(b.dataset.v);
  render();
}
function sortBy(col){
  if(sC===col)sD=-sD;else{sC=col;sD=1;}
  document.querySelectorAll('thead th').forEach(th=>th.classList.remove('sort-asc','sort-desc'));
  document.querySelectorAll('.sort-ic').forEach(ic=>ic.textContent='\u2195');
  const ic=document.getElementById('si-'+col);
  if(ic){ic.textContent=sD>0?'\u2191':'\u2193';ic.closest('th').classList.add(sD>0?'sort-asc':'sort-desc');}
  render();
}
const esc=v=>String(v||'').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');
function renderRow(r){
  let c='';
  const kp=(r.join_key||'').split('\u00a7');
  KEY_FIELDS.forEach((_,i)=>{c+='<td class="tk">'+esc(kp[i]||'')+'</td>';});
  const badges=r.ecarts.map(e=>{
    const st=BADGE_STYLE[e.rule_type]||BADGE_STYLE['ko'];
    let tip='';
    if(e.source_field||e.target_field)
      tip=esc(e.source_field)+'\u00a0=\u00a0'+esc(e.source_value)+' / '+esc(e.target_field)+'\u00a0=\u00a0'+esc(e.target_value);
    else if(e.detail)tip=esc(e.detail);
    return '<span class="badge" style="'+st+'"'+(tip?' title="'+tip+'"':'')+'>'+esc(e.rule_name)+'</span>';
  }).join(' ');
  c+='<td><div class="td-ecarts">'+badges+'</div></td>';
  COL_ORDER.forEach(({side,col})=>{
    const v=((side==='ref'?r._ref:r._tgt)||{})[col]||'';
    c+='<td class="tv xc-'+side+'" title="'+esc(v)+'">'+esc(v)+'</td>';
  });
  return '<tr>'+c+'</tr>';
}
function render(){
  const q=(document.getElementById('srch')?.value||'').toLowerCase();
  let rows=ALL.filter(r=>{
    if(!r.ecarts.some(e=>aR.has(e.rule_name)))return false;
    if(q){
      const kv=[r.join_key,...r.ecarts.flatMap(e=>[e.rule_name,e.source_value,e.target_value,e.detail])];
      const xv=COL_ORDER.map(({side,col})=>((side==='ref'?r._ref:r._tgt)||{})[col]||'');
      if(![...kv,...xv].some(v=>String(v||'').toLowerCase().includes(q)))return false;
    }
    return true;
  });
  if(sC){
    const ki=sC.startsWith('key_')?parseInt(sC.slice(4)):-1;
    rows=[...rows].sort((a,b)=>{
      let va,vb;
      if(ki>=0){const pa=(a.join_key||'').split('\u00a7');const pb=(b.join_key||'').split('\u00a7');va=pa[ki]||'';vb=pb[ki]||'';}
      else{const {side,col}=COL_ORDER.find(e=>e.side+':'+e.col===sC)||{};va=side?((side==='ref'?a._ref:a._tgt)||{})[col]||'':'';vb=side?((side==='ref'?b._ref:b._tgt)||{})[col]||'':'';}
      return va.localeCompare(vb,'fr',{numeric:true})*sD;
    });
  }
  const tb=document.getElementById('tbody');
  const em=document.getElementById('empty');
  const sh=document.getElementById('shown-count');
  if(sh)sh.textContent=rows.length.toLocaleString('fr-FR')+' ligne'+(rows.length!==1?'s':'')+' affich\u00e9es';
  if(!rows.length){tb.innerHTML='';em.style.display='block';return;}
  em.style.display='none';
  tb.innerHTML=rows.map(renderRow).join('');
}
initCounts();render();
"""

_HTML_CSS = """
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Segoe UI',system-ui,sans-serif;background:#f1f5f9;color:#1e293b;font-size:14px}
.layout{display:flex;flex-direction:column;height:100vh;overflow:hidden}
header{padding:.6rem 1.25rem;background:#fff;border-bottom:1px solid #e2e8f0;display:flex;align-items:center;gap:.75rem;flex-wrap:wrap;flex-shrink:0}
h1{font-size:1rem;font-weight:700;letter-spacing:-.01em;flex:1;min-width:0;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.meta{font-size:.72rem;color:#64748b;white-space:nowrap}
.desc{font-size:.75rem;color:#475569;flex-basis:100%;margin-top:.1rem}
.summary-bar{padding:.5rem 1.25rem;background:#fff;border-bottom:1px solid #e2e8f0;display:flex;align-items:center;gap:.65rem;flex-wrap:wrap;font-size:.78rem;flex-shrink:0}
.sum-ref{color:#16a34a;font-weight:500}
.sum-tgt{color:#ea580c;font-weight:500}
.sum-ko{color:#dc2626;font-weight:500}
.sum-ok{color:#16a34a;font-weight:500}
.sum-shown{color:#64748b;font-weight:500;margin-left:auto}
.sum-sep{color:#e2e8f0}
.cards{display:flex;gap:.5rem;flex-wrap:wrap;padding:.5rem 1.25rem;background:#fff;border-bottom:1px solid #e2e8f0;flex-shrink:0}
.card{background:#f8fafc;border:1px solid #e2e8f0;border-radius:6px;padding:.35rem .8rem;text-align:center;min-width:76px}
.card .v{font-size:1.1rem;font-weight:700;font-variant-numeric:tabular-nums}
.card .l{font-size:.6rem;color:#94a3b8;margin-top:.1rem;text-transform:uppercase;letter-spacing:.04em}
.card.ca .v{color:#dc2626}.card.cb .v{color:#ea580c}.card.cd .v{color:#ca8a04}.card.co .v{color:#16a34a}
.card.cr-ko .v{color:#b45309}.card.cr-ok .v{color:#0369a1}
.filter-bar{padding:.4rem 1rem;background:#fff;border-bottom:1px solid #e2e8f0;display:flex;flex-wrap:wrap;gap:.3rem .4rem;align-items:center;flex-shrink:0}
.chip{background:none;border:1px solid #cbd5e0;color:#94a3b8;font-size:.66rem;padding:.16rem .5rem;border-radius:99px;cursor:pointer;font-family:inherit;transition:background .12s,color .12s,border-color .12s;display:inline-flex;align-items:center;gap:.3rem}
.chip:hover{color:#475569;border-color:#94a3b8}
.chip.on.orphelin-source{background:#fef2f2;border-color:#fca5a5;color:#dc2626}
.chip.on.orphelin-cible{background:#fff7ed;border-color:#fdba74;color:#ea580c}
.chip.on.key-matched{background:#f0fdf4;border-color:#86efac;color:#16a34a}
.chip.on.cr-coh{background:#f0fdf4;border-color:#86efac;color:#16a34a}
.chip.on.cr-inc{background:#fef2f2;border-color:#fca5a5;color:#dc2626}
.chip-c{background:rgba(0,0,0,.07);border-radius:99px;padding:0 .32rem;font-size:.58rem;font-variant-numeric:tabular-nums}
.fl{font-size:.62rem;color:#94a3b8;flex-shrink:0}
.filter-sep{width:1px;height:14px;background:#e2e8f0;margin:0 .15rem;flex-shrink:0;align-self:center}
#srch{border:1px solid #e2e8f0;border-radius:4px;padding:.2rem .5rem;font-size:.72rem;color:#1e293b;background:#f8fafc;outline:none;width:180px;margin-left:auto}
#srch:focus{border-color:#6366f1;background:#fff}
.tbl-wrap{flex:1;overflow:auto}
table{width:100%;border-collapse:collapse;font-size:.75rem;background:#fff}
thead th{background:#fff;padding:.42rem .75rem;text-align:left;font-size:.6rem;font-weight:600;text-transform:uppercase;letter-spacing:.06em;color:#94a3b8;border-bottom:2px solid #e2e8f0;position:sticky;top:0;cursor:pointer;user-select:none;white-space:nowrap;z-index:1}
thead th:hover{background:#f8fafc;color:#475569}
.th-extra{white-space:normal;min-width:100px}
.th-meta{font-weight:300;font-size:.56rem;opacity:.7;text-transform:none;letter-spacing:0}
.th-field{font-weight:700;font-size:.63rem}
.th-ref .th-field{color:#1d4ed8}.th-tgt .th-field{color:#7c3aed}
tbody tr{border-bottom:1px solid #f1f5f9}
tbody tr:hover{background:#f8fafc}
td{padding:.38rem .75rem;vertical-align:middle}
.tk{font-family:'Cascadia Code','Fira Code',monospace;font-size:.7rem;color:#334155;white-space:nowrap}
.tv{font-family:'Cascadia Code','Fira Code',monospace;font-size:.7rem;max-width:160px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.tv.r{color:#1d4ed8}.tv.t{color:#c2410c}
.td-det{font-size:.66rem;color:#94a3b8;max-width:200px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.td-ecarts{display:flex;flex-wrap:wrap;gap:.25rem;align-items:center}
.badge{display:inline-block;padding:.1rem .4rem;border-radius:4px;font-size:.61rem;font-weight:700;white-space:nowrap;cursor:default}
.badge-_ko{background:#fee2e2;color:#b91c1c;border:1px solid #fecaca}
.badge-ko{background:#fef9c3;color:#a16207;border:1px solid #fef08a}
.badge-ok{background:#dcfce7;color:#15803d;border:1px solid #bbf7d0}
.badge-_ok{background:#eff6ff;color:#1d4ed8;border:1px solid #bfdbfe}
.sort-ic{opacity:.2;margin-left:.25rem;font-size:.55rem}
th.sort-asc .sort-ic,th.sort-desc .sort-ic{opacity:1;color:#6366f1}
.empty{text-align:center;padding:3rem;color:#cbd5e0;font-style:italic;display:none;font-size:.8rem}
"""


def _h(s) -> str:
    return str(s or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")


def to_html(results: list, summary: dict, config: dict,
            extra_ref: list = None, extra_tgt: list = None,
            ref_rows_map: dict = None, tgt_rows_map: dict = None,
            ref_label: str = "R\u00e9f\u00e9rence", tgt_label: str = "Cible",
            ref_fmt: str = "", tgt_fmt: str = "",
            extra_col_order: list = None) -> str:
    """HTML auto-suffisant — une ligne par clé, badges de règles, ordre de colonnes fidèle à la vue."""
    extra_ref    = extra_ref or []
    extra_tgt    = extra_tgt or []
    ref_rows_map = ref_rows_map or {}
    tgt_rows_map = tgt_rows_map or {}

    # Ordre mixte des colonnes supplémentaires (respecte l'ordre UI si fourni)
    if extra_col_order:
        col_order = [e for e in extra_col_order
                     if e.get("side") in ("ref", "tgt") and e.get("col")]
    else:
        col_order = (
            [{"side": "ref", "col": c} for c in extra_ref] +
            [{"side": "tgt", "col": c} for c in extra_tgt]
        )

    meta_cfg    = config.get("meta", {})
    name        = meta_cfg.get("name", "Audit")
    description = meta_cfg.get("description", "")
    now         = datetime.now().strftime("%d/%m/%Y \u00e0 %H:%M:%S")
    rules       = config.get("rules", [])
    cfg_keys    = config.get("join", {}).get("keys", [])
    key_fields  = [k.get("source_field", "Cl\u00e9") for k in cfg_keys] if cfg_keys else ["Cl\u00e9"]

    # Group flat results by join_key — one row per key with multiple ecarts
    grouped: dict = {}
    for r in results:
        key = r.get("join_key", "")
        if key not in grouped:
            grouped[key] = {
                "join_key": key,
                "ecarts":   [],
                "_ref": {c: ref_rows_map.get(key, {}).get(c, "") for c in extra_ref},
                "_tgt": {c: tgt_rows_map.get(key, {}).get(c, "") for c in extra_tgt},
            }
        grouped[key]["ecarts"].append({
            "rule_name":    r.get("rule_name",    ""),
            "rule_type":    r.get("rule_type",    ""),
            "source_field": r.get("source_field", ""),
            "target_field": r.get("target_field", ""),
            "source_value": r.get("source_value", ""),
            "target_value": r.get("target_value", ""),
            "detail":       r.get("detail",       ""),
        })
    grouped_list = list(grouped.values())

    # Rule chips
    rule_chips = ""
    if rules:
        rule_chips = '<div class="filter-sep"></div><span class="fl">R\u00e8gles</span>'
        for r_cfg in rules:
            rname = r_cfg.get("name", "")
            rtype = r_cfg.get("rule_type", "incoherence")
            cls   = "cr-coh" if rtype == "coherence" else "cr-inc"
            rule_chips += (
                f'<button class="chip {cls} on" data-k="rule" data-v="{_h(rname)}"'
                f' onclick="toggleChip(this)">{_h(rname)}'
                f' <span class="chip-c">0</span></button>'
            )

    # Per-rule KO cards from rule_stats
    rule_stats = summary.get("rule_stats", {})
    rule_cards = ""
    for r_cfg in rules:
        rname = r_cfg.get("name", "")
        n_ko  = rule_stats.get(rname, 0)
        rtype = r_cfg.get("rule_type", "incoherence")
        cls   = "cr-ok" if rtype == "coherence" else "cr-ko"
        lbl   = "OK" if rtype == "coherence" else "KO"
        rule_cards += (
            f'<div class="card {cls}">'
            f'<div class="v">{n_ko}</div>'
            f'<div class="l">{_h(rname)} \u2014 {lbl}</div>'
            f'</div>'
        )

    # Extra column TH elements (in col_order)
    ref_meta = _h(f"{ref_label} \u00b7 {ref_fmt}") if ref_fmt else _h(ref_label)
    tgt_meta = _h(f"{tgt_label} \u00b7 {tgt_fmt}") if tgt_fmt else _h(tgt_label)

    def _extra_th(side, col):
        meta  = ref_meta if side == "ref" else tgt_meta
        cls   = "th-ref" if side == "ref" else "th-tgt"
        sort_id = f"{side}:{_h(col)}"
        return (
            f'<th class="th-extra {cls}" onclick="sortBy(\'{sort_id}\')">'
            f'<div class="th-meta">{meta}</div>'
            f'<div class="th-field">{_h(col)}<span class="sort-ic" id="si-{sort_id}">\u2195</span></div></th>'
        )

    extra_ths = "".join(_extra_th(e["side"], e["col"]) for e in col_order)

    all_js        = json.dumps(grouped_list, ensure_ascii=False, default=str)
    col_order_js  = json.dumps(col_order)
    key_fields_js = json.dumps(key_fields)

    js = (
        _HTML_JS
        .replace("__ALL__",       all_js)
        .replace("__COL_ORDER__", col_order_js)
        .replace("__KEY_FIELDS__", key_fields_js)
    )

    s     = summary
    oa    = s.get("orphelins_a", 0)
    ob    = s.get("orphelins_b", 0)
    n_ko  = s.get("divergents",  0)
    n_ok  = s.get("ok",          0)
    n_tot = len(grouped_list)

    parts = [
        '<!DOCTYPE html>\n<html lang="fr"><head>'
        '<meta charset="UTF-8"><meta name="viewport" content="width=device-width">\n',
        f'<title>Rapport \u2014 {_h(name)}</title>\n',
        '<style>', _HTML_CSS, '</style></head><body><div class="layout">\n',

        # Header
        '<header>',
        f'<h1>{_h(name)}</h1>',
        f'<span class="meta">G\u00e9n\u00e9r\u00e9 le {_h(now)}</span>',
        *([ f'<span class="desc">{_h(description)}</span>' ] if description else []),
        '</header>\n',

        # Summary bar
        '<div class="summary-bar">',
        f'<span class="sum-ref">{_h(ref_label)}\u00a0: {s.get("total_reference",0):,} enr. dont {oa:,} absents de la cible</span>',
        '<span class="sum-sep">|</span>',
        f'<span class="sum-tgt">{_h(tgt_label)}\u00a0: {s.get("total_cible",0):,} enr. dont {ob:,} absents de la source</span>',
        '<span class="sum-sep">|</span>',
        f'<span class="sum-ko">{n_ko:,} KO</span>',
        '<span class="sum-sep">|</span>',
        f'<span class="sum-ok">{n_ok:,} OK</span>',
        f'<span class="sum-shown" id="shown-count">{n_tot:,} ligne{"s" if n_tot != 1 else ""} affich\u00e9es</span>',
        '</div>\n',

        # Per-rule cards
        *([ f'<div class="cards">{rule_cards}</div>\n' ] if rule_cards else []),

        # Filter bar
        '<div class="filter-bar">',
        '<span class="fl">Pr\u00e9sence</span>',
        f'<button class="chip orphelin-source on" data-k="rule" data-v="Source uniq." onclick="toggleChip(this)">'
        f'{_h(ref_label)} uniq. <span class="chip-c">0</span></button>',
        f'<button class="chip orphelin-cible on" data-k="rule" data-v="Cible uniq." onclick="toggleChip(this)">'
        f'{_h(tgt_label)} uniq. <span class="chip-c">0</span></button>',
        '<button class="chip key-matched on" data-k="rule" data-v="Cl\u00e9 OK" onclick="toggleChip(this)">'
        'Cl\u00e9 OK <span class="chip-c">0</span></button>',
        rule_chips,
        '<input id="srch" type="search" placeholder="Recherche\u2026" oninput="render()">',
        '</div>\n',

        # Table — column order: [keys] | [Règles badges] | [extra cols in col_order]
        '<div class="tbl-wrap"><table>\n<thead><tr>',
        "".join(
            f'<th onclick="sortBy(\'key_{i}\')">{_h(kf)}'
            f'<span class="sort-ic" id="si-key_{i}">\u2195</span></th>'
            for i, kf in enumerate(key_fields)
        ),
        '<th>R\u00e8gles</th>',
        extra_ths,
        '</tr></thead>\n',
        '<tbody id="tbody"></tbody>\n',
        '</table><div class="empty" id="empty">Aucun r\u00e9sultat pour les filtres s\u00e9lectionn\u00e9s.</div></div>',
        '</div>\n',
        '<script>', js, '</script>',
        '</body></html>',
    ]
    return "".join(parts)
