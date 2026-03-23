#!/usr/bin/env python3
"""
generate_test_data.py — Génère les jeux de données de test DataAuditor.

Produit dans sample/jdd/ :
  Référence : ref_long.csv, ref_long.txt, ref_wide.csv, ref_wide.txt,
              ref.json, ref.jsonl, ref.xlsx
  Cible     : tgt_long.csv, tgt_long.txt, tgt_wide.csv, tgt_wide.txt,
              tgt.json, tgt.jsonl, tgt.xlsx
  Configs   : cfg_csv.yaml, cfg_txt.yaml, cfg_wide_csv.yaml, cfg_wide_txt.yaml,
              cfg_json.yaml, cfg_jsonl.yaml, cfg_xlsx.yaml

Données : 20 produits × 3 sites (PAR/LYO/MAR) = 60 lignes utiles
Anomalies injectées côté cible (voir ANOMALIES ci-dessous).
Parasites : 4 produits × site ZZZ (filtrés par le filtre de source).
Pré-filtre TXT : lignes données préfixées "01", méta préfixée "00"/"99".
"""

import csv
import json
import os
import textwrap

OUT = os.path.join(os.path.dirname(__file__), "jdd")
os.makedirs(OUT, exist_ok=True)

# ─────────────────────────────────────────────
#  DONNÉES DE BASE
# ─────────────────────────────────────────────

# (ref, designation, categorie, prix_ht, date_inv, statut, ean13)
PRODUCTS = [
    ("P001", "Widget Alpha",         "A", 12.50, "20240115", "ACTIF",   "3012345000010"),
    ("P002", "Connecteur RJ45",      "B",  2.99, "20240120", "ACTIF",   "3012345000020"),
    ("P003", "Cable USB-C Pro",      "B",  8.99, "20240301", "ACTIF",   "3012345000030"),
    ("P004", "Support ecran 27p",    "A", 35.00, "20240210", "ACTIF",   "3012345000040"),
    ("P005", "Souris ergonomique",   "A", 24.90, "20240205", "ACTIF",   "3012345000050"),
    ("P006", "Clavier sans fil",     "A", 49.99, "20240115", "ACTIF",   "3012345000060"),
    ("P007", "Hub USB 3.0",          "B", 18.50, "20240220", "ACTIF",   "3012345000070"),
    ("P008", "Webcam HD 1080p",      "A", 42.00, "20240310", "ACTIF",   "3012345000080"),
    ("P009", "Casque audio BT",      "A", 89.99, "20240115", "ACTIF",   "3012345000090"),
    ("P010", "Chargeur 65W USB-C",   "B", 14.99, "20240201", "ACTIF",   "3012345000100"),
    ("P011", "Tapis de souris XL",   "C",  9.99, "20240115", "ACTIF",   "3012345000110"),
    ("P012", "Cable HDMI 2m",        "B",  5.99, "20240310", "ACTIF",   "3012345000120"),
    ("P013", "Station d accueil",    "A", 79.99, "20240125", "ACTIF",   "3012345000130"),
    ("P014", "Ecran 24p Full HD",    "A",189.00, "20240201", "ACTIF",   "3012345000140"),
    ("P015", "Alimentation 65W",     "B", 29.99, "20240310", "INACTIF", "3012345000150"),
    ("P016", "Boitier externe SSD",  "B", 22.50, "20240115", "ACTIF",   "3012345000160"),
    ("P017", "Cable DisplayPort 2m", "B",  7.99, "20240220", "ACTIF",   "3012345000170"),
    ("P018", "Adaptateur USB-C VGA", "B",  4.50, "20240301", "ACTIF",   "3012345000180"),
    ("P019", "Support ordinateur",   "C", 15.00, "20240210", "ACTIF",   "3012345000190"),
    ("P020", "Repose-poignets gel",  "C",  6.99, "20240115", "ACTIF",   "3012345000200"),
]

SITES = ["PAR", "LYO", "MAR"]

# Quantités de base par produit et site (fixes pour reproductibilité)
BASE_QTY = {
    "P001": {"PAR": 150, "LYO": 200, "MAR": 120},
    "P002": {"PAR":  80, "LYO":  60, "MAR":  45},
    "P003": {"PAR": 150, "LYO": 200, "MAR": 120},
    "P004": {"PAR":  45, "LYO":  30, "MAR":  25},
    "P005": {"PAR":  90, "LYO": 110, "MAR":  75},
    "P006": {"PAR":  55, "LYO":  70, "MAR":  40},
    "P007": {"PAR": 120, "LYO": 100, "MAR":  80},
    "P008": {"PAR":  35, "LYO":  25, "MAR":  20},
    "P009": {"PAR":  75, "LYO":  50, "MAR":  60},
    "P010": {"PAR": 200, "LYO": 180, "MAR": 150},
    "P011": {"PAR": 300, "LYO": 250, "MAR": 200},
    "P012": {"PAR": 180, "LYO": 160, "MAR": 140},
    "P013": {"PAR":  20, "LYO":  15, "MAR":  10},
    "P014": {"PAR":   8, "LYO":   5, "MAR":   4},
    "P015": {"PAR":  60, "LYO":  45, "MAR":  30},
    "P016": {"PAR":  85, "LYO":  70, "MAR":  55},
    "P017": {"PAR": 140, "LYO": 120, "MAR": 100},
    "P018": {"PAR": 250, "LYO": 220, "MAR": 190},
    "P019": {"PAR":  40, "LYO":  35, "MAR":  28},
    "P020": {"PAR": 110, "LYO":  90, "MAR":  70},
}

# Produits parasites (filtrés par site != ZZZ)
PARASITES = [
    ("P021", "Article obsolete A",  "C",  1.00, "20230101", "INACTIF", "3099999990010"),
    ("P022", "Article obsolete B",  "C",  1.00, "20230101", "INACTIF", "3099999990020"),
    ("P023", "Article test",        "C",  0.01, "20230101", "INACTIF", "3099999990030"),
    ("P024", "Article hors perim",  "C",  2.50, "20230601", "INACTIF", "3099999990040"),
]
PARASITE_QTY = {"P021": 5, "P022": 3, "P023": 1, "P024": 8}

# ─────────────────────────────────────────────
#  ANOMALIES (différences cible vs référence)
# ─────────────────────────────────────────────
# Clé = (ref, site), valeur = dict des champs modifiés côté cible
# "ABSENT" → ligne absente de la cible (ORPHELIN_A)
# "EXTRA"  → ligne présente uniquement en cible (ORPHELIN_B) — traitée à part
ANOMALIES = {
    # P001 : ORPHELIN_A — tous sites absents de la cible
    ("P001", "PAR"): "ABSENT",
    ("P001", "LYO"): "ABSENT",
    ("P001", "MAR"): "ABSENT",
    # P002 : ORPHELIN_B — présent uniquement en cible (géré séparément)
    # P003/PAR : KO Quantité (ref=150, tgt=165)
    ("P003", "PAR"): {"qty": 165},
    # P004/LYO : KO Prix (ref=35.00, tgt=41.50 — hors tolérance 0.50)
    ("P004", "LYO"): {"prix_ht": 41.50},
    # P005/MAR : OK Prix (ref=24.90, tgt=24.93 — dans tolérance 0.05)
    ("P005", "MAR"): {"prix_ht": 24.93},
    # P006/PAR : KO Désignation (casse différente, sans normalisation)
    ("P006", "PAR"): {"designation": "CLAVIER SANS FIL"},
    # P007/LYO : KO Statut
    ("P007", "LYO"): {"statut": "INACTIF"},
    # P008/MAR : KO EAN13 (dernier chiffre modifié)
    ("P008", "MAR"): {"ean13": "3012345000089"},
    # P009/PAR : KO Quantité ET KO Prix simultanément (multi-badges)
    ("P009", "PAR"): {"qty": 90, "prix_ht": 99.99},
    # P015 : INACTIF dans les deux → OK (montré si show_matching: true)
    # Pas d'anomalie — statut "INACTIF" identique dans ref et tgt
}

# P002 est le produit ORPHELIN_B : absent de la référence, présent uniquement en cible.
# Il ne doit PAS apparaître dans la boucle PRODUCTS côté ref.
ORPHELIN_B_REF = "P002"

# ─────────────────────────────────────────────
#  CONSTRUCTION DES LIGNES
# ─────────────────────────────────────────────

def build_rows(side):
    """
    side = 'ref' ou 'tgt'
    Retourne une liste de dicts : ref, designation, categorie, site,
                                  qty, prix_ht, date_inventaire, statut, ean13
    """
    rows = []
    prod_map = {p[0]: p for p in PRODUCTS}

    for ref, designation, categorie, prix_ht, date_inv, statut, ean13 in PRODUCTS:
        # ORPHELIN_B : absent de la référence
        if side == "ref" and ref == ORPHELIN_B_REF:
            continue

        for site in SITES:
            qty = BASE_QTY[ref][site]
            row = {
                "ref": ref,
                "designation": designation,
                "categorie": categorie,
                "site": site,
                "qty": qty,
                "prix_ht": prix_ht,
                "date_inventaire": date_inv,
                "statut": statut,
                "ean13": ean13,
            }
            key = (ref, site)

            if side == "tgt":
                anom = ANOMALIES.get(key)
                if anom == "ABSENT":
                    continue   # ligne absente de la cible (ORPHELIN_A)
                if isinstance(anom, dict):
                    row.update(anom)

            rows.append(row)

    # Parasites (site ZZZ) — présents dans les deux sources
    for ref, designation, categorie, prix_ht, date_inv, statut, ean13 in PARASITES:
        qty = PARASITE_QTY[ref]
        rows.append({
            "ref": ref, "designation": designation, "categorie": categorie,
            "site": "ZZZ", "qty": qty,
            "prix_ht": prix_ht, "date_inventaire": date_inv,
            "statut": statut, "ean13": ean13,
        })

    return rows

# ─────────────────────────────────────────────
#  FORMAT LONG — CSV
# ─────────────────────────────────────────────
LONG_FIELDS = ["ref", "designation", "categorie", "site", "qty",
               "prix_ht", "date_inventaire", "statut", "ean13"]

def write_csv(path, rows):
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=LONG_FIELDS, delimiter=";",
                           extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)
    print(f"  ✓ {os.path.relpath(path)}")

# ─────────────────────────────────────────────
#  FORMAT LONG — TXT avec pré-filtre
# ─────────────────────────────────────────────
# Structure :
#   00 → ligne méta (en-tête fichier, ignorée par record_filter "^01")
#   01 → ligne de donnée
#   99 → pied de fichier (total lignes, ignoré)

def write_txt(path, rows, label):
    lines = []
    # En-tête méta (non filtré → exclu par record_filter)
    lines.append(f"00;EXPORT_{label};2024-03-15;{len(rows)} enr.")
    # Données
    for r in rows:
        parts = [
            "01",
            r["ref"], r["designation"], r["categorie"], r["site"],
            str(r["qty"]), str(r["prix_ht"]), r["date_inventaire"],
            r["statut"], r["ean13"],
        ]
        lines.append(";".join(parts))
    # Pied de fichier
    data_lines = [l for l in lines if l.startswith("01")]
    lines.append(f"99;TOTAL={len(data_lines)}")

    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")
    print(f"  ✓ {os.path.relpath(path)}")

# ─────────────────────────────────────────────
#  FORMAT LARGE / PIVOTÉ — CSV et TXT
# ─────────────────────────────────────────────
WIDE_BASE_FIELDS = ["ref", "designation", "categorie",
                    "prix_ht", "date_inventaire", "statut", "ean13"]
WIDE_FIELDS = WIDE_BASE_FIELDS + ["qty_PAR", "qty_LYO", "qty_MAR"]

def build_wide_rows(side):
    """Une ligne par produit. qty_PAR/LYO/MAR des lignes long du même side."""
    long_rows = build_rows(side)
    # Indexer par (ref, site)
    idx = {}
    for r in long_rows:
        idx[(r["ref"], r["site"])] = r

    prod_map = {p[0]: p for p in PRODUCTS}
    wide = {}

    # Produits de référence (on itère toujours sur PRODUCTS)
    for ref, designation, categorie, prix_ht, date_inv, statut, ean13 in PRODUCTS:
        # ORPHELIN_B : absent de la référence
        if side == "ref" and ref == ORPHELIN_B_REF:
            continue
        if side == "tgt" and ANOMALIES.get((ref, "PAR")) == "ABSENT":
            # P001 absent de la cible → pas de ligne large non plus
            skip_all = all(ANOMALIES.get((ref, s)) == "ABSENT" for s in SITES)
            if skip_all:
                continue
        row = {
            "ref": ref, "designation": designation, "categorie": categorie,
            "prix_ht": prix_ht, "date_inventaire": date_inv,
            "statut": statut, "ean13": ean13,
        }
        # Appliquer anomalies sur prix_ht / statut / designation
        for site in SITES:
            anom = ANOMALIES.get((ref, site))
            if isinstance(anom, dict):
                if "prix_ht" in anom and side == "tgt":
                    row["prix_ht"] = anom["prix_ht"]
                if "statut" in anom and side == "tgt":
                    row["statut"] = anom["statut"]
                if "designation" in anom and side == "tgt":
                    row["designation"] = anom["designation"]

        for site in SITES:
            key = (ref, site)
            r = idx.get(key)
            qty = r["qty"] if r else 0
            row[f"qty_{site}"] = qty

        wide[ref] = row

    # Orphelin B en cible
    if side == "tgt":
        ref, designation, categorie, prix_ht, date_inv, statut, ean13 = prod_map["P002"]
        wide[ref] = {
            "ref": ref, "designation": designation, "categorie": categorie,
            "prix_ht": prix_ht, "date_inventaire": date_inv,
            "statut": statut, "ean13": ean13,
            "qty_PAR": BASE_QTY[ref]["PAR"],
            "qty_LYO": BASE_QTY[ref]["LYO"],
            "qty_MAR": BASE_QTY[ref]["MAR"],
        }

    # Parasites (site ZZZ non pertinent en wide — on les exclut)
    return list(wide.values())

def write_wide_csv(path, rows):
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=WIDE_FIELDS, delimiter=";",
                           extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)
    print(f"  ✓ {os.path.relpath(path)}")

def write_wide_txt(path, rows, label):
    lines = []
    lines.append(f"00;EXPORT_WIDE_{label};2024-03-15;{len(rows)} enr.")
    for r in rows:
        parts = ["01", r["ref"], r["designation"], r["categorie"],
                 str(r["prix_ht"]), r["date_inventaire"], r["statut"], r["ean13"],
                 str(r["qty_PAR"]), str(r["qty_LYO"]), str(r["qty_MAR"])]
        lines.append(";".join(parts))
    data_lines = [l for l in lines if l.startswith("01")]
    lines.append(f"99;TOTAL={len(data_lines)}")
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")
    print(f"  ✓ {os.path.relpath(path)}")

# ─────────────────────────────────────────────
#  JSON (structure imbriquée)
# ─────────────────────────────────────────────
# Chaque enregistrement :
# { "ref", "designation", "categorie",
#   "localisation": {"site": ...},
#   "stock":        {"qty": ..., "date_inventaire": ...},
#   "commercial":   {"prix_ht": ..., "statut": ...},
#   "ean13" }
# json_path : export.records

def row_to_json_obj(r):
    return {
        "ref": r["ref"],
        "designation": r["designation"],
        "categorie": r["categorie"],
        "localisation": {"site": r["site"]},
        "stock": {
            "qty": r["qty"],
            "date_inventaire": r["date_inventaire"],
        },
        "commercial": {
            "prix_ht": float(r["prix_ht"]),
            "statut": r["statut"],
        },
        "ean13": r["ean13"],
    }

def write_json(path, rows, label):
    records = [row_to_json_obj(r) for r in rows]
    obj = {
        "export": {
            "source": label,
            "date": "2024-03-15",
            "records": records,
        }
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False, indent=2)
    print(f"  ✓ {os.path.relpath(path)}")

def write_jsonl(path, rows):
    with open(path, "w", encoding="utf-8") as f:
        for r in rows:
            f.write(json.dumps(row_to_json_obj(r), ensure_ascii=False) + "\n")
    print(f"  ✓ {os.path.relpath(path)}")

# ─────────────────────────────────────────────
#  XLSX
# ─────────────────────────────────────────────

def write_xlsx(path, rows):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print(f"  ⚠ openpyxl absent — {os.path.relpath(path)} non généré")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stock"

    headers = LONG_FIELDS
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, r in enumerate(rows, 2):
        for col_idx, field in enumerate(headers, 1):
            val = r[field]
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Largeurs automatiques
    col_widths = {"ref": 8, "designation": 25, "categorie": 10, "site": 6,
                  "qty": 8, "prix_ht": 10, "date_inventaire": 14,
                  "statut": 10, "ean13": 15}
    for col, h in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = col_widths.get(h, 12)

    wb.save(path)
    print(f"  ✓ {os.path.relpath(path)}")

# ─────────────────────────────────────────────
#  CONFIGS YAML
# ─────────────────────────────────────────────

RULES_YAML = """\
rules:
  - name: "Quantite"
    logic: AND
    rule_type: incoherence
    fields:
      - source_field: qty
        target_field: qty
        operator: differs
        tolerance: 0

  - name: "Prix HT"
    logic: AND
    rule_type: incoherence
    fields:
      - source_field: prix_ht
        target_field: prix_ht
        operator: differs
        tolerance: 0.05

  - name: "Designation"
    logic: AND
    rule_type: incoherence
    fields:
      - source_field: designation
        target_field: designation
        operator: differs

  - name: "Statut"
    logic: AND
    rule_type: incoherence
    fields:
      - source_field: statut
        target_field: statut
        operator: differs

  - name: "Code-barres"
    logic: AND
    rule_type: incoherence
    fields:
      - source_field: ean13
        target_field: ean13
        operator: differs
"""

FILTERS_YAML = """\
filters:
  - field: site
    source: reference
    operator: differs
    value: ZZZ
  - field: site
    source: target
    operator: differs
    value: ZZZ
"""

REPORT_YAML = """\
report:
  show_matching: true
  max_diff_preview: 500
"""

JOIN_LONG_YAML = """\
join:
  keys:
    - source_field: ref
      target_field: ref
    - source_field: site
      target_field: site
"""

JOIN_WIDE_YAML = """\
join:
  keys:
    - source_field: ref
      target_field: ref
    - source_field: site
      target_field: site
"""

UNPIVOT_REF_YAML = """\
    unpivot:
      anchor_fields: [ref, designation, categorie, prix_ht, date_inventaire, statut, ean13]
      location_field: site
      value_field: qty
      pivot_fields:
        - source: qty_PAR
          location: PAR
        - source: qty_LYO
          location: LYO
        - source: qty_MAR
          location: MAR
"""

def write_yaml(path, content):
    with open(path, "w", encoding="utf-8") as f:
        f.write(content)
    print(f"  ✓ {os.path.relpath(path)}")


def cfg_csv():
    return f"""\
meta:
  name: "Test CSV long format"
  version: "1.0"

sources:
  reference:
    label: "Reference WMS"
    format: csv
    encoding: utf-8
    delimiter: ";"
    has_header: true
    fields:
      - {{name: ref,             type: string }}
      - {{name: designation,     type: string }}
      - {{name: categorie,       type: string }}
      - {{name: site,            type: string }}
      - {{name: qty,             type: integer}}
      - {{name: prix_ht,         type: decimal}}
      - {{name: date_inventaire, type: date, date_format: "%Y%m%d"}}
      - {{name: statut,          type: string }}
      - {{name: ean13,           type: string }}

  target:
    label: "Cible ERP"
    format: csv
    encoding: utf-8
    delimiter: ";"
    has_header: true
    fields:
      - {{name: ref,             type: string }}
      - {{name: designation,     type: string }}
      - {{name: categorie,       type: string }}
      - {{name: site,            type: string }}
      - {{name: qty,             type: integer}}
      - {{name: prix_ht,         type: decimal}}
      - {{name: date_inventaire, type: date, date_format: "%Y%m%d"}}
      - {{name: statut,          type: string }}
      - {{name: ean13,           type: string }}

{JOIN_LONG_YAML}
{FILTERS_YAML}
{RULES_YAML}
{REPORT_YAML}"""


def cfg_txt():
    return f"""\
meta:
  name: "Test TXT avec pre-filtre"
  version: "1.0"

sources:
  reference:
    label: "Reference WMS"
    format: csv
    encoding: utf-8
    delimiter: ";"
    has_header: false
    record_filter:
      marker: "^01"
    fields:
      - {{name: type_enreg,     type: skip   }}
      - {{name: ref,             type: string }}
      - {{name: designation,     type: string }}
      - {{name: categorie,       type: string }}
      - {{name: site,            type: string }}
      - {{name: qty,             type: integer}}
      - {{name: prix_ht,         type: decimal}}
      - {{name: date_inventaire, type: date, date_format: "%Y%m%d"}}
      - {{name: statut,          type: string }}
      - {{name: ean13,           type: string }}

  target:
    label: "Cible ERP"
    format: csv
    encoding: utf-8
    delimiter: ";"
    has_header: false
    record_filter:
      marker: "^01"
    fields:
      - {{name: type_enreg,     type: skip   }}
      - {{name: ref,             type: string }}
      - {{name: designation,     type: string }}
      - {{name: categorie,       type: string }}
      - {{name: site,            type: string }}
      - {{name: qty,             type: integer}}
      - {{name: prix_ht,         type: decimal}}
      - {{name: date_inventaire, type: date, date_format: "%Y%m%d"}}
      - {{name: statut,          type: string }}
      - {{name: ean13,           type: string }}

{JOIN_LONG_YAML}
{FILTERS_YAML}
{RULES_YAML}
{REPORT_YAML}"""


def cfg_wide_csv():
    return f"""\
meta:
  name: "Test CSV pivoté (wide → long via unpivot)"
  version: "1.0"

sources:
  reference:
    label: "Reference WMS (wide)"
    format: csv
    encoding: utf-8
    delimiter: ";"
    has_header: true
    fields:
      - {{name: ref,             type: string }}
      - {{name: designation,     type: string }}
      - {{name: categorie,       type: string }}
      - {{name: prix_ht,         type: decimal}}
      - {{name: date_inventaire, type: date, date_format: "%Y%m%d"}}
      - {{name: statut,          type: string }}
      - {{name: ean13,           type: string }}
      - {{name: qty_PAR,         type: integer}}
      - {{name: qty_LYO,         type: integer}}
      - {{name: qty_MAR,         type: integer}}
{UNPIVOT_REF_YAML}
  target:
    label: "Cible ERP (long)"
    format: csv
    encoding: utf-8
    delimiter: ";"
    has_header: true
    fields:
      - {{name: ref,             type: string }}
      - {{name: designation,     type: string }}
      - {{name: categorie,       type: string }}
      - {{name: site,            type: string }}
      - {{name: qty,             type: integer}}
      - {{name: prix_ht,         type: decimal}}
      - {{name: date_inventaire, type: date, date_format: "%Y%m%d"}}
      - {{name: statut,          type: string }}
      - {{name: ean13,           type: string }}

{JOIN_WIDE_YAML}
{FILTERS_YAML}
{RULES_YAML}
{REPORT_YAML}"""


def cfg_wide_txt():
    return f"""\
meta:
  name: "Test TXT pivoté avec pre-filtre"
  version: "1.0"

sources:
  reference:
    label: "Reference WMS (wide TXT)"
    format: csv
    encoding: utf-8
    delimiter: ";"
    has_header: false
    record_filter:
      marker: "^01"
    fields:
      - {{name: type_enreg,     type: skip   }}
      - {{name: ref,             type: string }}
      - {{name: designation,     type: string }}
      - {{name: categorie,       type: string }}
      - {{name: prix_ht,         type: decimal}}
      - {{name: date_inventaire, type: date, date_format: "%Y%m%d"}}
      - {{name: statut,          type: string }}
      - {{name: ean13,           type: string }}
      - {{name: qty_PAR,         type: integer}}
      - {{name: qty_LYO,         type: integer}}
      - {{name: qty_MAR,         type: integer}}
{UNPIVOT_REF_YAML}
  target:
    label: "Cible ERP (long TXT)"
    format: csv
    encoding: utf-8
    delimiter: ";"
    has_header: false
    record_filter:
      marker: "^01"
    fields:
      - {{name: type_enreg,     type: skip   }}
      - {{name: ref,             type: string }}
      - {{name: designation,     type: string }}
      - {{name: categorie,       type: string }}
      - {{name: site,            type: string }}
      - {{name: qty,             type: integer}}
      - {{name: prix_ht,         type: decimal}}
      - {{name: date_inventaire, type: date, date_format: "%Y%m%d"}}
      - {{name: statut,          type: string }}
      - {{name: ean13,           type: string }}

{JOIN_WIDE_YAML}
{FILTERS_YAML}
{RULES_YAML}
{REPORT_YAML}"""


def cfg_json():
    return f"""\
meta:
  name: "Test JSON imbriqué"
  version: "1.0"

sources:
  reference:
    label: "Reference WMS"
    format: json
    encoding: utf-8
    json_path: export.records
    fields:
      - {{name: ref,             type: string }}
      - {{name: designation,     type: string }}
      - {{name: categorie,       type: string }}
      - {{name: site,            type: string,  path: localisation.site}}
      - {{name: qty,             type: integer, path: stock.qty}}
      - {{name: date_inventaire, type: date, date_format: "%Y%m%d", path: stock.date_inventaire}}
      - {{name: prix_ht,         type: decimal, path: commercial.prix_ht}}
      - {{name: statut,          type: string,  path: commercial.statut}}
      - {{name: ean13,           type: string }}

  target:
    label: "Cible ERP"
    format: json
    encoding: utf-8
    json_path: export.records
    fields:
      - {{name: ref,             type: string }}
      - {{name: designation,     type: string }}
      - {{name: categorie,       type: string }}
      - {{name: site,            type: string,  path: localisation.site}}
      - {{name: qty,             type: integer, path: stock.qty}}
      - {{name: date_inventaire, type: date, date_format: "%Y%m%d", path: stock.date_inventaire}}
      - {{name: prix_ht,         type: decimal, path: commercial.prix_ht}}
      - {{name: statut,          type: string,  path: commercial.statut}}
      - {{name: ean13,           type: string }}

{JOIN_LONG_YAML}
{FILTERS_YAML}
{RULES_YAML}
{REPORT_YAML}"""


def cfg_jsonl():
    return f"""\
meta:
  name: "Test JSONL imbriqué"
  version: "1.0"

sources:
  reference:
    label: "Reference WMS"
    format: jsonl
    encoding: utf-8
    fields:
      - {{name: ref,             type: string }}
      - {{name: designation,     type: string }}
      - {{name: categorie,       type: string }}
      - {{name: site,            type: string,  path: localisation.site}}
      - {{name: qty,             type: integer, path: stock.qty}}
      - {{name: date_inventaire, type: date, date_format: "%Y%m%d", path: stock.date_inventaire}}
      - {{name: prix_ht,         type: decimal, path: commercial.prix_ht}}
      - {{name: statut,          type: string,  path: commercial.statut}}
      - {{name: ean13,           type: string }}

  target:
    label: "Cible ERP"
    format: jsonl
    encoding: utf-8
    fields:
      - {{name: ref,             type: string }}
      - {{name: designation,     type: string }}
      - {{name: categorie,       type: string }}
      - {{name: site,            type: string,  path: localisation.site}}
      - {{name: qty,             type: integer, path: stock.qty}}
      - {{name: date_inventaire, type: date, date_format: "%Y%m%d", path: stock.date_inventaire}}
      - {{name: prix_ht,         type: decimal, path: commercial.prix_ht}}
      - {{name: statut,          type: string,  path: commercial.statut}}
      - {{name: ean13,           type: string }}

{JOIN_LONG_YAML}
{FILTERS_YAML}
{RULES_YAML}
{REPORT_YAML}"""


def cfg_xlsx():
    return f"""\
meta:
  name: "Test XLSX"
  version: "1.0"

sources:
  reference:
    label: "Reference WMS"
    format: xlsx
    sheet_name: Stock

  target:
    label: "Cible ERP"
    format: xlsx
    sheet_name: Stock

{JOIN_LONG_YAML}
{FILTERS_YAML}
{RULES_YAML}
{REPORT_YAML}"""


# ─────────────────────────────────────────────
#  SOURCE BLOCKS (réutilisables en cross-format)
# ─────────────────────────────────────────────

SRC_CSV = """\
    format: csv
    encoding: utf-8
    delimiter: ";"
    has_header: true
    fields:
      - {name: ref,             type: string }
      - {name: designation,     type: string }
      - {name: categorie,       type: string }
      - {name: site,            type: string }
      - {name: qty,             type: integer}
      - {name: prix_ht,         type: decimal}
      - {name: date_inventaire, type: date, date_format: "%Y%m%d"}
      - {name: statut,          type: string }
      - {name: ean13,           type: string }"""

SRC_TXT = """\
    format: csv
    encoding: utf-8
    delimiter: ";"
    has_header: false
    record_filter:
      marker: "^01"
    fields:
      - {name: type_enreg,     type: skip   }
      - {name: ref,             type: string }
      - {name: designation,     type: string }
      - {name: categorie,       type: string }
      - {name: site,            type: string }
      - {name: qty,             type: integer}
      - {name: prix_ht,         type: decimal}
      - {name: date_inventaire, type: date, date_format: "%Y%m%d"}
      - {name: statut,          type: string }
      - {name: ean13,           type: string }"""

SRC_JSON = """\
    format: json
    encoding: utf-8
    json_path: export.records
    fields:
      - {name: ref,             type: string }
      - {name: designation,     type: string }
      - {name: categorie,       type: string }
      - {name: site,            type: string,  path: localisation.site}
      - {name: qty,             type: integer, path: stock.qty}
      - {name: date_inventaire, type: date, date_format: "%Y%m%d", path: stock.date_inventaire}
      - {name: prix_ht,         type: decimal, path: commercial.prix_ht}
      - {name: statut,          type: string,  path: commercial.statut}
      - {name: ean13,           type: string }"""

SRC_JSONL = """\
    format: jsonl
    encoding: utf-8
    fields:
      - {name: ref,             type: string }
      - {name: designation,     type: string }
      - {name: categorie,       type: string }
      - {name: site,            type: string,  path: localisation.site}
      - {name: qty,             type: integer, path: stock.qty}
      - {name: date_inventaire, type: date, date_format: "%Y%m%d", path: stock.date_inventaire}
      - {name: prix_ht,         type: decimal, path: commercial.prix_ht}
      - {name: statut,          type: string,  path: commercial.statut}
      - {name: ean13,           type: string }"""

SRC_XLSX = """\
    format: xlsx
    sheet_name: Stock"""


def cfg_cross(name, ref_label, tgt_label, ref_src, tgt_src):
    return f"""\
meta:
  name: "{name}"
  version: "1.0"

sources:
  reference:
    label: "{ref_label}"
{ref_src}

  target:
    label: "{tgt_label}"
{tgt_src}

{JOIN_LONG_YAML}
{FILTERS_YAML}
{RULES_YAML}
{REPORT_YAML}"""


# ─────────────────────────────────────────────
#  GÉNÉRATION
# ─────────────────────────────────────────────

def main():
    print("\n=== Génération des jeux de données de test ===\n")

    ref_long = build_rows("ref")
    tgt_long = build_rows("tgt")
    ref_wide = build_wide_rows("ref")
    tgt_wide = build_wide_rows("tgt")

    print(f"Référence long : {len(ref_long)} lignes")
    print(f"Cible long     : {len(tgt_long)} lignes")
    print(f"Référence wide : {len(ref_wide)} lignes")
    print(f"Cible wide     : {len(tgt_wide)} lignes\n")

    # CSV long
    print("--- CSV long ---")
    write_csv(f"{OUT}/ref_long.csv", ref_long)
    write_csv(f"{OUT}/tgt_long.csv", tgt_long)

    # TXT long
    print("--- TXT long (avec pré-filtre) ---")
    write_txt(f"{OUT}/ref_long.txt", ref_long, "REF")
    write_txt(f"{OUT}/tgt_long.txt", tgt_long, "TGT")

    # CSV wide
    print("--- CSV wide (pivoté) ---")
    write_wide_csv(f"{OUT}/ref_wide.csv", ref_wide)
    write_wide_csv(f"{OUT}/tgt_wide.csv", tgt_wide)

    # TXT wide
    print("--- TXT wide (pivoté + pré-filtre) ---")
    write_wide_txt(f"{OUT}/ref_wide.txt", ref_wide, "REF")
    write_wide_txt(f"{OUT}/tgt_wide.txt", tgt_wide, "TGT")

    # JSON
    print("--- JSON (imbriqué) ---")
    write_json(f"{OUT}/ref.json", ref_long, "WMS")
    write_json(f"{OUT}/tgt.json", tgt_long, "ERP")

    # JSONL
    print("--- JSONL (imbriqué) ---")
    write_jsonl(f"{OUT}/ref.jsonl", ref_long)
    write_jsonl(f"{OUT}/tgt.jsonl", tgt_long)

    # XLSX
    print("--- XLSX ---")
    write_xlsx(f"{OUT}/ref.xlsx", ref_long)
    write_xlsx(f"{OUT}/tgt.xlsx", tgt_long)

    # YAML configs — same format
    print("--- Configurations YAML (même format) ---")
    write_yaml(f"{OUT}/cfg_csv.yaml",      cfg_csv())
    write_yaml(f"{OUT}/cfg_txt.yaml",      cfg_txt())
    write_yaml(f"{OUT}/cfg_wide_csv.yaml", cfg_wide_csv())
    write_yaml(f"{OUT}/cfg_wide_txt.yaml", cfg_wide_txt())
    write_yaml(f"{OUT}/cfg_json.yaml",     cfg_json())
    write_yaml(f"{OUT}/cfg_jsonl.yaml",    cfg_jsonl())
    write_yaml(f"{OUT}/cfg_xlsx.yaml",     cfg_xlsx())

    # YAML configs — cross-format
    print("--- Configurations YAML (cross-format) ---")
    write_yaml(f"{OUT}/cfg_csv_vs_jsonl.yaml", cfg_cross(
        "CSV → JSONL (cross-format)",
        "Reference WMS (CSV)", "Cible ERP (JSONL)",
        SRC_CSV, SRC_JSONL,
    ))
    write_yaml(f"{OUT}/cfg_txt_vs_csv.yaml", cfg_cross(
        "TXT positionnel → CSV (cross-format)",
        "Reference WMS (TXT positionnel)", "Cible ERP (CSV)",
        SRC_TXT, SRC_CSV,
    ))
    write_yaml(f"{OUT}/cfg_json_vs_csv.yaml", cfg_cross(
        "JSON imbriqué → CSV (cross-format)",
        "Reference WMS (JSON)", "Cible ERP (CSV)",
        SRC_JSON, SRC_CSV,
    ))
    write_yaml(f"{OUT}/cfg_xlsx_vs_jsonl.yaml", cfg_cross(
        "XLSX → JSONL (cross-format)",
        "Reference WMS (XLSX)", "Cible ERP (JSONL)",
        SRC_XLSX, SRC_JSONL,
    ))

    print("\n=== Résumé des anomalies injectées ===")
    print("  ORPHELIN_A : P001 (toutes sites manquants en cible)")
    print("  ORPHELIN_B : P002 (toutes sites présents uniquement en cible)")
    print("  KO Quantite          : P003/PAR (ref=150, tgt=165)")
    print("  KO Prix HT           : P004/LYO (ref=35.00, tgt=41.50, hors tol=0.05)")
    print("  OK Prix HT           : P005/MAR (ref=24.90, tgt=24.93, dans tol=0.05)")
    print("  KO Designation       : P006/PAR (ref='Clavier sans fil', tgt='CLAVIER SANS FIL')")
    print("  KO Statut            : P007/LYO (ref=ACTIF, tgt=INACTIF)")
    print("  KO Code-barres       : P008/MAR (ean dernier chiffre modifie)")
    print("  KO Quantite + Prix   : P009/PAR (multi-regles simultanees)")
    print("  Parasites filtres    : P021-P024 site ZZZ")
    print("  Pre-filtre TXT       : lignes '00' (meta) et '99' (total) ignorees")
    print("\nDone.\n")


if __name__ == "__main__":
    main()
